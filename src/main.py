"""
Automação de Votação — versão otimizada
"""

import logging
import os
import re
import unicodedata
from collections import defaultdict
from pathlib import Path
from typing import Optional

from openpyxl import load_workbook
from rapidfuzz import fuzz

# =========================
#  CONFIGURAÇÃO
# =========================
logging.basicConfig(
    level=logging.INFO,
    format="%(levelname)s | %(message)s",
)
log = logging.getLogger(__name__)

FUZZY_THRESHOLD = 85
MIN_PALAVRAS_SUBSET = 2
COL_INICIO = 12  # coluna L
COL_STATUS = "I"
COL_NOME = "D"

ASSESSORES_LEGAIS: dict[str, str] = {
    "PGA": "Pinheiro Guimarães Advogados & CSW Advogados",
    "MM":  "Machado Meyer Advogados",
    "FEL": "Felsberg Advogados",
    "CTP": "Costa Tavares Paes Advogados",
}

ASSESSORES_FINANCEIROS: dict[str, str] = {
    "BR":   "BR partners",
    "G5":   "G5 partners consultoria",
    "JNEY": "Journey Capital",
    "VIR":  "Virtus",
}

MAPA_VOTOS: dict[str, str] = {
    "A":  "sim",
    "R":  "não",
    "AB": "ab",
    "NV": "nv",
    "não": "não",
    "sim": "sim",
}

entrada = "SIM"
entrada = "NÃO"

resultado = MAPA_VOTOS.get(entrada.lower())


# =========================
#  NORMALIZAÇÃO
# =========================
def normalizar_nome(nome: str) -> str:
    """Remove acentos, caracteres especiais e normaliza espaços."""
    nome = str(nome).upper()
    nome = unicodedata.normalize("NFKD", nome)
    nome = "".join(c for c in nome if not unicodedata.combining(c))
    nome = re.sub(r"[^A-Z0-9 ]", " ", nome)
    return " ".join(nome.split())



# =========================
#  LEITURA DE ARQUIVOS
# =========================
def listar_arquivos_empresa(base_path: Path, empresa: str) -> list[Path]:
    caminho = base_path / empresa
    arquivos = [f for f in caminho.iterdir() if f.is_file()]
    if not arquivos:
        log.warning("Pasta vazia: %s", empresa)
    return arquivos


def extrair_dados(caminho_arquivo: Path) -> tuple[Optional[str], list[str]]:
    """
    Extrai nome normalizado e lista de votos a partir do nome do arquivo.
    Retorna (None, []) em caso de formato inválido.
    """
    stem = caminho_arquivo.stem.replace('"', "")
    partes = re.split(r"\s*[-–—]\s*", stem, maxsplit=1)

    if len(partes) < 2:
        log.warning("Nome fora do padrão: %s", caminho_arquivo.name)
        return None, []

    nome = partes[0].strip()
    votos_str = partes[1].rstrip(",").strip()
    votos = [v.strip() for v in votos_str.split(";") if v.strip()]
    return normalizar_nome(nome), votos


# =========================
#  PROCESSAMENTO
# =========================
def processar_dados(base_path: Path) -> dict[str, list[dict]]:
    mapa_dados: dict[str, list[dict]] = defaultdict(list)

    empresas = [p.name for p in base_path.iterdir() if p.is_dir()]
    if not empresas:
        log.warning("Nenhuma pasta de empresa encontrada em: %s", base_path)
        return mapa_dados

    for empresa in empresas:
        for arquivo in listar_arquivos_empresa(base_path, empresa):
            nome, votos = extrair_dados(arquivo)
            if not nome:
                continue
            mapa_dados[nome].append({
                "empresa": empresa,
                "votos":   votos,
                "arquivo": str(arquivo),
            })

    duplicados = {n: r for n, r in mapa_dados.items() if len(r) > 1}
    if duplicados:
        log.warning("Duplicados encontrados:")
        for nome, registros in duplicados.items():
            log.warning("  %s — %d arquivos", nome, len(registros))

    return mapa_dados


# =========================
#  MATCH DE NOMES
# =========================
def encontrar_nome_aproximado(
    nome_planilha: str,
    mapa_dados: dict[str, list[dict]],
    usados: set[str],
) -> Optional[str]:
    """
    Tenta casar nome_planilha com uma chave em mapa_dados.
    Prioridade:
      1. Subconjunto exato: todas as palavras sig. do arquivo estão na planilha
      2. Interseção alta: palavras que batem cobrem >= 80% de ambos os lados
      3. Fuzzy matching acima de FUZZY_THRESHOLD
    """
    STOPWORDS = {"DE", "DA", "DO", "DAS", "DOS", "E", "EM", "A", "O", "AS", "OS"}

    def palavras_significativas(nome: str) -> set[str]:
        return {p for p in nome.split() if p not in STOPWORDS and len(p) > 2}

    palavras_sig_planilha = palavras_significativas(nome_planilha)
    melhor_match: Optional[str] = None
    melhor_score = 0

    for nome_arquivo, _ in mapa_dados.items():
        if nome_arquivo in usados:
            continue

        palavras_sig_arquivo = palavras_significativas(nome_arquivo)

        if len(palavras_sig_arquivo) < MIN_PALAVRAS_SUBSET:
            continue

        intersecao = palavras_sig_arquivo & palavras_sig_planilha

        # Regra 1: subconjunto exato (arquivo ⊆ planilha)
        if palavras_sig_arquivo.issubset(palavras_sig_planilha):
            usados.add(nome_arquivo)
            return nome_arquivo

        # Regra 2: todas as palavras da planilha estão no arquivo (planilha é subconjunto do arquivo)
        # Cobre casos como "Erick Yamada" (planilha) vs "Erick dos Santos Yamada" (arquivo)
        if palavras_sig_planilha.issubset(palavras_sig_arquivo):
            usados.add(nome_arquivo)
            return nome_arquivo

        # Regra 3: fuzzy sobre palavras significativas
        score = fuzz.token_sort_ratio(
            " ".join(sorted(palavras_sig_planilha)),
            " ".join(sorted(palavras_sig_arquivo)),
        )
        if score > melhor_score:
            melhor_score = score
            melhor_match = nome_arquivo

    if melhor_match and melhor_score >= FUZZY_THRESHOLD:
        usados.add(melhor_match)
        return melhor_match

    return None


def traduzir_voto(valor: str) -> str:
    """Converte código de voto para texto legível."""
    valor = re.sub(r"[^A-Z0-9]", "", valor.strip().upper())  # remove vírgulas, pontos, etc.
    if valor in MAPA_VOTOS:
        return MAPA_VOTOS[valor]
    if valor in ASSESSORES_LEGAIS:
        return ASSESSORES_LEGAIS[valor]
    if valor in ASSESSORES_FINANCEIROS:
        return ASSESSORES_FINANCEIROS[valor]
    return valor


# =========================
#  EXCEL
# =========================
def detectar_colunas_itens(ws, col_inicio: int) -> int:
    """Conta colunas de itens a partir de col_inicio até None ou 'SÉRIE'."""
    col = col_inicio
    while True:
        valor = ws.cell(row=1, column=col).value
        if valor is None:
            break
        if isinstance(valor, str) and "SÉRIE" in valor.upper():
            break
        col += 1
    total = col - col_inicio
    log.info("Colunas de itens detectadas: %d", total)
    return total


def escrever_excel(
    excel_path: Path,
    mapa_dados: dict[str, list[dict]],
) -> tuple[Path, dict[str, Optional[str]]]:
    """
    Preenche o Excel e retorna (caminho_salvo, resultado_matches).

    resultado_matches: { nome_arquivo -> nome_planilha matched (ou None se não encontrou) }
    Arquivos que já estavam com 'ok' não entram no resultado (já foram processados antes).
    """
    wb = load_workbook(excel_path)
    ws = wb["COMITENTES"]

    max_colunas_itens = detectar_colunas_itens(ws, COL_INICIO)
    usados: set[str] = set()

    # resultado_matches: chave = nome_arquivo, valor = nome_planilha que fez match (ou None)
    resultado_matches: dict[str, Optional[str]] = {}

    linha = 2
    encontrados = nao_encontrados = pulados = 0

    while True:
        nome_planilha = ws[f"{COL_NOME}{linha}"].value
        if nome_planilha is None:
            break

        nome_norm = normalizar_nome(nome_planilha)

        # Se a linha já tem status "ok", marca o arquivo como usado para a auditoria
        status_atual = ws[f"{COL_STATUS}{linha}"].value
        if status_atual == "ok":
            log.info("Já preenchido, pulando: %s", nome_planilha)
            # Registra que esse nome da planilha já estava resolvido,
            # marcando o arquivo correspondente como usado (se encontrar)
            match = encontrar_nome_aproximado(nome_norm, mapa_dados, usados)
            if match:
                resultado_matches[match] = nome_planilha
            pulados += 1
            linha += 1
            continue

        nome_match = encontrar_nome_aproximado(nome_norm, mapa_dados, usados)

        if nome_match:
            votos = mapa_dados[nome_match][0]["votos"]
            ws[f"{COL_STATUS}{linha}"] = "ok"
            for i in range(min(max_colunas_itens, len(votos))):
                celula = ws.cell(row=linha, column=COL_INICIO + i)
                if celula.value is None:
                    celula.value = traduzir_voto(votos[i])
            resultado_matches[nome_match] = nome_planilha
            encontrados += 1
        else:
            log.info("Sem match: %s", nome_planilha)
            nao_encontrados += 1

        linha += 1

    wb.save(excel_path)

    log.info("Planilha salva: %s", excel_path)
    log.info(
        "Adicionados: %d | Já existiam (pulados): %d | Sem match: %d",
        encontrados, pulados, nao_encontrados,
    )
    return excel_path, resultado_matches


# =========================
#  DIAGNÓSTICO
# =========================
def auditar_pendentes(
    mapa_dados: dict[str, list[dict]],
    resultado_matches: dict[str, Optional[str]],
) -> None:
    """
    Identifica arquivos da pasta que não foram associados a nenhuma linha com 'ok'.
    Usa o resultado já calculado por escrever_excel — sem refazer nenhum match.
    """
    arquivos_matched = set(resultado_matches.keys())
    todos_arquivos   = set(mapa_dados.keys())
    sem_ok           = todos_arquivos - arquivos_matched

    print("\n=== AUDITORIA DE PENDENTES ===")
    if not sem_ok:
        print("✓ Todos os arquivos da pasta já têm 'ok' na planilha.")
        return

    print(f"{len(sem_ok)} arquivo(s) sem 'ok':\n")
    for nome_arquivo in sorted(sem_ok):
        registros = mapa_dados[nome_arquivo]
        caminhos  = ", ".join(r["arquivo"] for r in registros)
        print(f"  ARQUIVO : {nome_arquivo}")
        print(f"  CAMINHO : {caminhos}")
        print()


# =========================
#  MAIN
# =========================
def main() -> None:
    print("=== Automação de Votação ===\n")

    base_path = Path(input("Caminho da pasta com arquivos: ").strip())
    excel_path = Path(input("Caminho do arquivo Excel: ").strip())

    if not base_path.is_dir():
        log.error("Caminho da pasta inválido: %s", base_path)
        return

    if not excel_path.is_file():
        log.error("Caminho do Excel inválido: %s", excel_path)
        return

    mapa_dados = processar_dados(base_path)

    modo = input("\nO que deseja fazer?\n  1 - Preencher planilha\n  2 - Auditar pendentes\n  3 - Ambos\nOpção: ").strip()

    # escrever_excel sempre roda para manter o set de usados completo,
    # inclusive para linhas já com "ok" — isso garante que a auditoria
    # receba o mapa de matches correto sem refazer nenhuma lógica.
    _, resultado_matches = escrever_excel(excel_path, mapa_dados)

    if modo in ("2", "3"):
        auditar_pendentes(mapa_dados, resultado_matches)


if __name__ == "__main__":
    main()